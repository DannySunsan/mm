import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

class excelProxy:
    __wb = None;
    __index = 0
    __font = None
    __pattern = None
    __border = None
    __alignment = None
    def __init__(self):
        self.__wb = openpyxl.Workbook()

        self.setFont()
        self.setAlignment()
        self.setPattern() 
        self.setBorder() 
    def addSheet(self,name:str):
        if self.__index == 0:            
            self.__wb.active.title = name
            self.__index +=  1
            return
        self.__wb.active = self.__wb.create_sheet(name,self.__index)
        self.__index +=  1

    def sheetNum(self):
        return self.__index

    def setValue(self,r,c,value):
        cell = self.__wb.active.cell(r,c,value)
        cell.font = self.__font
        cell.alignment = self.__alignment
        cell.border = self.__border
        cell.fill = self.__pattern

    def merge(self,rs,re,cs,ce):
        self.__wb.active.merge_cells(None,rs,cs,re,ce)
        
    def setFont(self,name='宋体',size=10,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'):
        self.__font = Font(name, size, bold, italic, vertAlign, 
                                    underline, strike,color)

    def setPattern(self,fill_type='darkUp',start_color='FFFFFF',end_color='FFFFFF'):
        self.__pattern = PatternFill(fill_type, start_color, end_color)

    def setBorder(self,left=Side(border_style='medium',color='FF0000'),
                right=Side(border_style='medium',color='FF0000'),
                top=Side(border_style='medium',color='000000'),
                bottom=Side(border_style='medium',color='0000FF')):
        self.__border = Border(left,right,top,bottom)

    def setAlignment(self,horizontal='left',
                    vertical='center',
                    text_rotation=0,
                    indent=0):
        self.__alignment = Alignment(horizontal,vertical,text_rotation,indent)

    #Excel 默认列宽为 8.38
    def setColWidth(self,index:str,width:int):        
        self.__wb.active.column_dimensions[index].width = width
    #Excel 默认行高为 13.5
    def setRowHeight(self,index:int,height:int):
        if index < 1:
            return
        self.__wb.active.rown_dimensions[index].height = height

    def save(self,path:str):
        self.__wb.save(path)

